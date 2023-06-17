import streamlit as st
import pandas as pd
from io import StringIO
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os 

import pandas as pd
import numpy as np

import os
from pathlib import Path

from glob import glob

import shutil

width_list = [10.5,8.88,7.75,7.25,6.5,51.88]
width_list_op = [8.38,8.88,7.75,7.25,56.88,51.88]

def convert_df(df):
    # IMPORTANT: Cache the conversion to prevent computation on every rerun
    return df.to_csv().encode('utf-8')

def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'}) 
    worksheet.set_column('A:A', None, format1)  
    writer.close()
    processed_data = output.getvalue()
    return processed_data

adm_xlsx = st.file_uploader("입원목록")
dc_xlsx = st.file_uploader("퇴원목록")
op_xlsx = st.file_uploader("수술목록")

writer = pd.ExcelWriter('Total.xlsx', engine='xlsxwriter')
writer.close()

if adm_xlsx is not None : 
    adm = pd.read_excel(adm_xlsx, header = 2)
    adm = adm[['입원일자','환자번호','성  명','Sex/Age','담당교수','진단명']]
    writer = pd.ExcelWriter('입원.xlsx', engine='xlsxwriter')
    writer_total = pd.ExcelWriter('Total.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay')
    adm_date_list = adm.입원일자.unique().tolist()
    for adm_date in adm_date_list : 
        df_date = adm[adm.입원일자 == adm_date]
        df_date.to_excel(writer, sheet_name=f'{adm_date} - 입원')  
        df_date.to_excel(writer_total, sheet_name=f'{adm_date} - 입원')  
        
    writer.close()
    writer_total.save()
    
    wb_a = load_workbook("입원.xlsx")
    wb_t = load_workbook("Total.xlsx")
    
    for adm_date in adm_date_list : 
        ws_a = wb_a[f'{adm_date} - 입원']
        ws_a.column_dimensions['B'].width = width_list[0]
        ws_a.column_dimensions['C'].width  = width_list[1]
        ws_a.column_dimensions['D'].width  = width_list[2]
        ws_a.column_dimensions['E'].width  = width_list[3]
        ws_a.column_dimensions['F'].width  = width_list[4]
        ws_a.column_dimensions['G'].width  = width_list[5]
        for i in range(1,50):
            ai = ws_a[f'G{i}']
            ai.alignment = Alignment(horizontal='left', vertical='top',wrap_text = True) 
            
        ws_t = wb_t[f'{adm_date} - 입원']
        ws_t.column_dimensions['B'].width = width_list[0]
        ws_t.column_dimensions['C'].width  = width_list[1]
        ws_t.column_dimensions['D'].width  = width_list[2]
        ws_t.column_dimensions['E'].width  = width_list[3]
        ws_t.column_dimensions['F'].width  = width_list[4]
        ws_t.column_dimensions['G'].width  = width_list[5]
        for i in range(1,50):
            ai = ws_t[f'G{i}']
            ai.alignment = Alignment(horizontal='left', vertical='top',wrap_text = True) 
         

    wb_a.save('입원.xlsx')
    wb_t.save('Total.xlsx')
    
    df_adm = pd.read_excel('입원.xlsx')
    with open("입원.xlsx", "rb") as file:
        btn = st.download_button(
                label="입원",
                data=file,
                file_name="입원.xlsx",
                mime="application/vnd.ms-excel"
              )

if dc_xlsx is not None : 
    dc = pd.read_excel(dc_xlsx , header = 2)
    dc = dc[['퇴원일자','환자번호','성  명','Sex/Age','담당교수','진단명']]
    writer = pd.ExcelWriter('퇴원.xlsx', engine='xlsxwriter')
    writer_total = pd.ExcelWriter('Total.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay')
    dc_date_list = dc.퇴원일자.unique().tolist()
    for dc_date in sorted(dc_date_list) : 
        df_date = dc[dc.퇴원일자 == dc_date]
        df_date.to_excel(writer, sheet_name=f'{dc_date} - 퇴원')
        df_date.to_excel(writer_total, sheet_name=f'{dc_date} - 퇴원')
        
    writer.close()
    writer_total.save()
    wb_d = load_workbook("퇴원.xlsx")
    wb_t = load_workbook("Total.xlsx")
    
    for dc_date in sorted(dc_date_list) : 
        ws_d = wb_d[f'{dc_date} - 퇴원']
        ws_d.column_dimensions['B'].width = width_list[0]
        ws_d.column_dimensions['C'].width  = width_list[1]
        ws_d.column_dimensions['D'].width  = width_list[2]
        ws_d.column_dimensions['E'].width  = width_list[3]
        ws_d.column_dimensions['F'].width  = width_list[4]
        ws_d.column_dimensions['G'].width  = width_list[5]
        for i in range(1,50):
            ai = ws_d[f'G{i}']
            ai.alignment = Alignment(horizontal='left', vertical='top',wrap_text = True) 
            
        ws_t = wb_t[f'{dc_date} - 퇴원']
        ws_t.column_dimensions['B'].width = width_list[0]
        ws_t.column_dimensions['C'].width  = width_list[1]
        ws_t.column_dimensions['D'].width  = width_list[2]
        ws_t.column_dimensions['E'].width  = width_list[3]
        ws_t.column_dimensions['F'].width  = width_list[4]
        ws_t.column_dimensions['G'].width  = width_list[5]
        for i in range(1,50):
            ai = ws_t[f'G{i}']
            ai.alignment = Alignment(horizontal='left', vertical='top',wrap_text = True) 

    wb_d.save('퇴원.xlsx')
    wb_t.save('Total.xlsx') 
    
    df_dc = pd.read_excel('퇴원.xlsx')
    with open("퇴원.xlsx", "rb") as file:
        btn = st.download_button(
                label="퇴원",
                data=file,
                file_name="퇴원.xlsx",
                mime="application/vnd.ms-excel"
              )


if op_xlsx is not None : 

    op = pd.read_excel(op_xlsx , header = 2)
    op = op[['수술일자','등록번호','성명','성별/나이','수술명','진단명']]
    writer = pd.ExcelWriter('수술.xlsx', engine='xlsxwriter')
    writer_total = pd.ExcelWriter('Total.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay')
    op_date_list = op.수술일자.unique().tolist()
    for op_date in sorted(op_date_list) : 
        df_date = op[op.수술일자 == op_date]
        df_date.to_excel(writer, sheet_name=f'20{op_date} - 수술')    
        df_date.to_excel(writer_total, sheet_name=f'20{op_date} - 수술')  
    writer.close()
    writer_total.save()
    
    wb_o = load_workbook("수술.xlsx")
    wb_t = load_workbook("Total.xlsx")
    for op_date in sorted(op_date_list) : 
        ws_o = wb_o[f'20{op_date} - 수술']
        ws_o.column_dimensions['B'].width = width_list_op[0]
        ws_o.column_dimensions['C'].width  = width_list_op[1]
        ws_o.column_dimensions['D'].width  = width_list_op[2]
        ws_o.column_dimensions['E'].width  = width_list_op[3]
        ws_o.column_dimensions['F'].width  = width_list_op[4]
        ws_o.column_dimensions['G'].width  = width_list_op[5]
        for i in range(1,50):
            ai = ws_o[f'G{i}']
            ai.alignment = Alignment(horizontal='left', vertical='top',wrap_text = True) 
            ai = ws_o[f'F{i}']
            ai.alignment = Alignment(horizontal='left', vertical='top',wrap_text = True) 
            
        ws_t = wb_t[f'20{op_date} - 수술']
        ws_t.column_dimensions['B'].width = width_list_op[0]
        ws_t.column_dimensions['C'].width  = width_list_op[1]
        ws_t.column_dimensions['D'].width  = width_list_op[2]
        ws_t.column_dimensions['E'].width  = width_list_op[3]
        ws_t.column_dimensions['F'].width  = width_list_op[4]
        ws_t.column_dimensions['G'].width  = width_list_op[5]
        for i in range(1,50):
            ai = ws_t[f'G{i}']
            ai.alignment = Alignment(horizontal='left', vertical='top',wrap_text = True) 
            ai = ws_t[f'F{i}']
            ai.alignment = Alignment(horizontal='left', vertical='top',wrap_text = True) 
    wb_t._sheets.sort(key=lambda ws:ws.title)
    wb_o.save('수술.xlsx')
    wb_t.save('Total.xlsx')
    
    df_op = pd.read_excel('수술.xlsx')
    with open("수술.xlsx", "rb") as file:
        btn = st.download_button(
                label="수술",
                data=file,
                file_name="수술.xlsx",
                mime="application/vnd.ms-excel"
              )

if (op_xlsx is not None) & (adm_xlsx is not None) & (dc_xlsx is not None) : 
    df_total = pd.read_excel('Total.xlsx')
    with open("Total.xlsx", "rb") as file:
        btn = st.download_button(
                label="Total",
                data=file,
                file_name="Total.xlsx",
                mime="application/vnd.ms-excel"
              )